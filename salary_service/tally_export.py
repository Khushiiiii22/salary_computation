from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter


def export_to_tally_excel(df, output_path):
    """
    Converts the processed salary DataFrame into Excel format
    compatible with Tally import. Tally often expects specific columns
    like Ledger Name, Amount, Particulars, etc. Adjust per Tally requirement.

    Improvements:
    - Header styling (bold, center aligned)
    - Currency number formatting for monetary columns
    - Auto width for columns
    - Input validation
    - Basic error handling
    """
    required_columns = ['Employee_ID', 'Employee_Name', 'Net_Salary', 'Total_Deductions', 'Gross_Salary']

    # Validate input DataFrame columns
    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Input DataFrame missing required columns: {missing_cols}")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Salary Voucher"

        headers = ['Employee_ID', 'Employee_Name', 'Amount (Net Salary)', 'Deductions', 'Gross Salary']
        ws.append(headers)

        # Style headers: bold and center aligned
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = center_alignment

        # Write data rows with rounding
        for idx, row in df.iterrows():
            ws.append([
                row['Employee_ID'],
                row['Employee_Name'],
                round(row['Net_Salary'], 2),
                round(row['Total_Deductions'], 2),
                round(row['Gross_Salary'], 2)
            ])

        # Apply currency number format to salary columns
        currency_format = numbers.BUILTIN_FORMATS[7]  # Format like $#,##0.00; - see Excel built-in formats

        # Assuming columns C (3), D (4), E (5) have monetary values
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=5):
            for cell in row:
                cell.number_format = currency_format

        # Auto-adjust column widths based on max content length
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    cell_value = str(cell.value)
                    if cell_value:
                        max_length = max(max_length, len(cell_value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output_path)

    except Exception as e:
        raise RuntimeError(f"Failed to export to Excel: {e}")
